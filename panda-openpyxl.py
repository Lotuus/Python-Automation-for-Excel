from openpyxl import Workbook
import pandas as pd

gdir = r'C:\Users\izhan\OneDrive\Desktop\Python Files\atom-projects\Math-IA'  # directory of Math-IA; use ('{}\\ ...').format()

wb = Workbook()
st = wb.active

flnm = r'C:\Users\Isaac\Desktop\Python Files\atom-projects\ia.xlsx'  # original file to reference off of
df = pd.read_excel(flnm)

# store columns into lists
dat = list(df['Date'])
nor = list(df['Normal'])
d0 = list(df['D0'])
d1 = list(df['D1'])
d2 = list(df['D2'])
d3 = list(df['D3'])
d4 = list(df['D4'])


# analyze chunks, input appropriate details for avg, Poisson distribution, etc to sheet
def anch(column, size, criteria):
    cl = column  # column to analyze
    s = size  # number of rows in a chunk
    cr = criteria  # criteria for analyzing chunks
    lcntr = []  # list to store counters from analyzing chunks

    # step through column list in chunks, adds counter of indices in the chunk that meet the criteria to a list
    for i in range(0, len(cl), s):
        chunk = cl[i:i+s]
        cntr = 0

        if len(chunk) < s:
            lcntr.append('END')  # if there is a chunk that does not meet intended size, skip this chunk
        else:
            for j in chunk:
                if j > cr:
                    cntr += 1
            lcntr.append(cntr)

    # assign sheet column header names and column number corresponding to the column
    if cl is nor:
        clno = 1
        h = 'Normal'
    elif cl is d0:
        clno = 2
        h = 'D0'
    elif cl is d1:
        clno = 3
        h = 'D1'
    elif cl is d2:
        clno = 4
        h = 'D2'
    elif cl is d3:
        clno = 5
        h = 'D3'
    elif cl is d4:
        clno = 6
        h = 'D4'

    st.cell(row=1, column=clno).value = h

    # input chunk criteria counter list into cells at the appropriate column
    for r in range(0, len(lcntr)):  # if starting range is eg. 2, there will be 2 chunks missing; thus, range should start at 0
        st.cell(row=r+2, column=clno).value = lcntr[r]  # start one row below header

    # input average and its header into sheet
    st.cell(row=1, column=9).value = 'AVG'
    if lcntr[len(lcntr)-1] == 'END':  # if a chunk has been skipped due to size, avg calculation will reflect that
        avg = sum(lcntr[0:len(lcntr)-1])/(len(lcntr)-1)
    else:
        avg = sum(lcntr[0:len(lcntr)-1])/len(lcntr)

    st.cell(row=2, column=9).value = avg  # avg for calculations in sheet formula keeps full precision

    # input probability headers and perform Poisson PDF calculations
    st.cell(row=4, column=9).value = 'Pois({})'.format('%s' % float('%.6g' % avg))  # header name rounded to 6 sig fig

    for r in range(0, int(s*1.5)+1):
        st.cell(row=r+5, column=9).value = '=POISSON.DIST({}, $I$2, FALSE)'.format(r)  # =POISSON.DIST(I5, $I$2, FALSE)

    # assign file name corresponding to column, chunk size, and criteria, and save the file
    if h == 'Normal':
        flnm = '{}_{}w_{}.xlsx'.format('Normal', s, cr)
    else:
        flnm = '{}_{}w_{}.xlsx'.format(h, s, cr)

    wb.save(filename=flnm)


anch(d4, 4, 1)
anch(d4, 24, 1)
anch(d4, 52, 1)

# TODO the string formatting seems to omitt 0s at the end of the avg when formatting to sig figsg
