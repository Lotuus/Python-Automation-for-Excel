from openpyxl import load_workbook

# other sheet opening options
# <read_only> loads a workbook in read-only mode, which saves time when loading large files
# <data_only> loads a workbook without its formulas, only its values

xlsx = 'ia.xlsx'  # variable for file name to work on
lwb = load_workbook(filename='ia.xlsx')  # workbook to read; can instead link full directory address
print('Selected workbook:', xlsx)
print('All available sheet(s):', lwb.sheetnames)  # prints name of all available sheets in seleceted workbook

st = lwb.active
print('Selected sheet:', st.title)  # prints name of the selected sheet, the first one by default
# can select other sheets by name: |wb.get_sheet_by_name('<name>')|

# to print a single cell, not a selection
# print(st['A1'].value)  # prints the selected cell

# to iterate through a selection of cells
# note that <.value> cannot print `tuple` selections, as the -tuple- object has no =value= attribute
# for val in st.iter_rows(min_row=1, max_row=10, min_col=2, max_col=2, values_only=True):  # prints values in selection
#     print(val)

# for rw in st.rows:  # prints -cell- objects for selection, in this case, all rows in selected sheet
#     print(rw)

col = 2  # column to work at
rwmx = st.max_row  # total row count in sheet
rwmn = st.min_row  # lowest ith row in sheet
rws = 1  # first row to iterate at
rwl = 11  # rows to add onto to first grabbed row for grouping; WATCH FOR +1 AT rwmx+1
rwe = rws + rwl  # ending row
