from openpyxl import load_workbook

# other sheet opening options
# <read_only> loads a workbook in read-only mode, which saves time when loading large files
# <data_only> loads a workbook without its formulas, only its values

lwb = load_workbook(filename='hello-world.xlsx')  # workbook to read
print('All available sheet(s):', lwb.sheetnames)  # prints name of all available sheets in selected workbook

st = lwb.active
print('Selected sheet:', st.title)  # prints selected sheet, the first one by default

# to call a cell, but not access its content
# print(st["A1"])  # :<Cell 'testtest'.A1>
# print(st.cell(row=10, column=6))  # :<Cell 'testtest'.F10>; this is a -cell- object

# to call and access a cell's content
print(st["A1"].value)
print(st.cell(row=2, column=1).value)  # note that Excel indices begin at 1 instead of 0
